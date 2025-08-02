# app.py - Phase 1: Setup + User Inputs + Nanoparticle Generation

import streamlit as st
import random
import os
from ase.cluster import FaceCenteredCubic, BodyCenteredCubic, HexagonalClosedPacked
from ase.io import write
import tempfile

# --- Constants ---
BULK_COORD = 12  # Coordination number to define surface atoms

# --- Page Config ---
st.set_page_config(
    page_title="Monte Carlo Nanoparticle Simulator - Phase 1",
    page_icon="⚛️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# --- Title ---
st.title("⚛️ Monte Carlo Nanoparticle Simulator — Phase 1: Setup & Nanoparticle Generation")

# --- Sidebar: User Input Section ---
with st.sidebar:
    st.header("Configure Simulation Parameters")

    element_A = st.text_input(
        "Element A Symbol",
        value="Pt",
        max_chars=2,
        help="Chemical symbol of element A (e.g., Pt, Au, Fe)"
    )

    element_B = st.text_input(
        "Element B Symbol",
        value="Ru",
        max_chars=2,
        help="Chemical symbol of element B (e.g., Ru, Sn, Co)"
    )

    composition_A = st.slider(
        "Composition of Element A",
        min_value=0.0,
        max_value=1.0,
        value=0.5,
        step=0.01,
        help="Fraction of element A atoms in the nanoparticle (0 to 1)"
    )

    temperature = st.number_input(
        "Simulation Temperature (K)",
        min_value=0.0,
        max_value=5000.0,
        value=250.0,
        step=1.0,
        help="Temperature in Kelvin for Monte Carlo simulation (used in later phases)"
    )

    n_steps = st.number_input(
        "Number of Monte Carlo Steps",
        min_value=100,
        max_value=100000,
        value=10000,
        step=100,
        help="Total Monte Carlo steps to simulate (used in later phases)"
    )

    save_interval = st.number_input(
        "Data Save Interval (steps)",
        min_value=10,
        max_value=10000,
        value=1000,
        step=10,
        help="Interval between saving log data (used in later phases)"
    )

    snapshot_interval = st.number_input(
        "Snapshot Save Interval (steps)",
        min_value=10,
        max_value=10000,
        value=500,
        step=10,
        help="Interval between saving nanoparticle snapshots as .xyz files"
    )

    lattice_type = st.selectbox(
        "Lattice Type",
        options=["fcc", "bcc", "hcp"],
        index=0,
        help="Crystal lattice type of the nanoparticle"
    )

    layers = st.slider(
        "Number of Layers (each dimension)",
        min_value=1,
        max_value=15,
        value=7,
        step=1,
        help="Number of atomic layers in x, y, z (cubic)"
    )

    # Optional: surfaces for cluster (fixed here as default FCC style)
    surfaces = [(1, 1, 1)] * 3  # can be extended later if needed

    st.markdown("---")
    st.caption("⚙️ Note: Temperature, steps, save intervals will be used in simulation phases.")

# --- Validate inputs ---
if element_A.strip() == "" or element_B.strip() == "":
    st.error("Please enter valid chemical symbols for both elements.")
    st.stop()

if composition_A < 0 or composition_A > 1:
    st.error("Composition must be between 0 and 1.")
    st.stop()

# --- Map lattice type to ASE cluster builder ---
lattice_map = {
    "fcc": FaceCenteredCubic,
    "bcc": BodyCenteredCubic,
    "hcp": HexagonalClosedPacked
}

ClusterBuilder = lattice_map.get(lattice_type.lower())
if ClusterBuilder is None:
    st.error(f"Unsupported lattice type '{lattice_type}'. Please select from fcc, bcc, hcp.")
    st.stop()

# --- Generate Nanoparticle Button ---
if st.button("🔹 Generate Nanoparticle"):

    with st.spinner("Generating nanoparticle structure..."):

        # Build cluster object
        particle = ClusterBuilder(
            element_A,
            layers=(layers, layers, layers),
            surfaces=surfaces,
        )

        n_atoms = len(particle)
        n_A = int(n_atoms * composition_A)

        # Randomly assign element A to the desired fraction of atoms
        indices_A = random.sample(range(n_atoms), n_A)
        for i in range(n_atoms):
            particle[i].symbol = element_A if i in indices_A else element_B

        # Save the initial structure to a temporary file
        temp_dir = tempfile.gettempdir()
        initial_xyz_path = os.path.join(temp_dir, f"{element_A}{element_B}_initial_{n_atoms}.xyz")
        write(initial_xyz_path, particle)

        # --- Show Results ---
        st.success(f"Nanoparticle generated with {n_atoms} atoms.")
        st.write(f"**Composition:** {element_A}: {composition_A*100:.2f}%, {element_B}: {(1-composition_A)*100:.2f}%")
        st.write(f"**Lattice type:** {lattice_type.upper()}")
        st.write(f"**Layers (x,y,z):** {layers} × {layers} × {layers}")

        # Show summary atom count table
        st.markdown("### Atomic Composition:")
        counts = {element_A: 0, element_B: 0}
        for atom in particle:
            counts[atom.symbol] = counts.get(atom.symbol, 0) + 1
        st.table({ "Element": list(counts.keys()), "Count": list(counts.values()) })

        # 3D Visualization using py3Dmol
        try:
            import py3Dmol

            with open(initial_xyz_path) as f:
                xyz_data = f.read()

            view = py3Dmol.view(width=600, height=500)
            view.addModel(xyz_data, "xyz")

            view.setStyle(
                { "elem": element_A },
                { "sphere": { "color": "gold", "radius": 1.2 } }
            )
            view.setStyle(
                { "elem": element_B },
                { "sphere": { "color": "green", "radius": 1.2 } }
            )

            view.setBackgroundColor("white")
            view.zoomTo()
            st.markdown("### 3D Visualization of Initial Nanoparticle")
            view.show()
        except ImportError:
            st.warning("`py3Dmol` package not installed. 3D visualization unavailable.")

        # Provide download of the XYZ file
        with open(initial_xyz_path, "rb") as file:
            st.download_button(
                label="📥 Download Initial Structure (XYZ file)",
                data=file,
                file_name=os.path.basename(initial_xyz_path),
                mime="chemical/x-xyz"
            )

        # Save particle & metadata to session state for later phases
        st.session_state['particle'] = particle
        st.session_state['initial_xyz'] = initial_xyz_path
        st.session_state['n_atoms'] = n_atoms
        st.session_state['element_A'] = element_A
        st.session_state['element_B'] = element_B
        st.session_state['composition_A'] = composition_A
        st.session_state['temperature'] = temperature
        st.session_state['n_steps'] = n_steps
        st.session_state['save_interval'] = save_interval
        st.session_state['snapshot_interval'] = snapshot_interval
        st.session_state['lattice_type'] = lattice_type

else:
    st.info("Click 'Generate Nanoparticle' to create your initial structure.")

# --- Footer ---
st.markdown("---")
st.caption("Developed by YourName | Monte Carlo Nanoparticle Simulation Project")
import math
import random
import time
import shutil
from ase.neighborlist import build_neighbor_list
from openpyxl import Workbook

# --- Helper Functions ---

def symbol_type(sym, element_A):
    return 'A' if sym == element_A else 'B'

def calculate_energy(particle, coeffs, element_A):
    nl = build_neighbor_list(particle, bothways=True, self_interaction=False)
    count = {k: 0 for k in coeffs.keys()}
    for i in range(len(particle)):
        t_i = symbol_type(particle[i].symbol, element_A)
        neighbors = nl.get_neighbors(i)[0]
        if len(neighbors) < BULK_COORD:
            count[f'x{t_i}-S'] += 1
        for j in neighbors:
            if i < j:
                t_j = symbol_type(particle[j].symbol, element_A)
                if t_i == t_j:
                    count[f'x{t_i}-{t_j}'] += 1
                else:
                    count['xA-B'] += 1
    energy = sum(coeffs[k] * count[k] for k in count)
    return energy

def count_surface(particle, element_A):
    nl = build_neighbor_list(particle, bothways=True, self_interaction=False)
    total_A, surf, surf_A = 0, 0, 0
    for i in range(len(particle)):
        neighbors = nl.get_neighbors(i)[0]
        t = symbol_type(particle[i].symbol, element_A)
        if t == 'A':
            total_A += 1
        if len(neighbors) < BULK_COORD:
            surf += 1
            if t == 'A':
                surf_A += 1
    ratio = surf_A / surf if surf > 0 else 0
    return total_A, surf, surf_A, ratio

# --- Phase 2 Main Monte Carlo Simulation Section ---
st.markdown("---")
st.header("⚛️ Phase 2: Monte Carlo Simulation & Data Logging")

if 'particle' not in st.session_state:
    st.warning("Please generate the nanoparticle structure first in Phase 1.")
    st.stop()

# Retrieve data from session state
particle = st.session_state['particle']
element_A = st.session_state['element_A']
element_B = st.session_state['element_B']
composition_A = st.session_state['composition_A']
temperature = st.session_state['temperature']
n_steps = st.session_state['n_steps']
save_interval = st.session_state['save_interval']
snapshot_interval = st.session_state['snapshot_interval']
lattice_type = st.session_state['lattice_type']
n_atoms = st.session_state['n_atoms']

# Default coefficients (You can later extend to user input)
coefficients = {
    'xA-A': -0.022078,
    'xB-B': -0.150000,
    'xA-B': -0.109575,
    'xA-S': -0.250717,
    'xB-S': -0.300000,
    'xA-A-out': 0.184150,
    'xB-B-out': 0.332228,
    'xA-B-out': 0.051042
}

# Create trajectory directory inside temp directory for snapshots
import tempfile
trajectory_dir = os.path.join(tempfile.gettempdir(), "trajectory")
os.makedirs(trajectory_dir, exist_ok=True)

# Run Monte Carlo on button click
if st.button("▶️ Run Monte Carlo Simulation"):

    energy = calculate_energy(particle, coefficients, element_A)
    log = []
    boltzmann_k = 8.617333262e-5  # eV/K

    progress_bar = st.progress(0)
    status_text = st.empty()
    start_time = time.time()

    for step in range(1, n_steps + 1):

        i = random.randint(0, n_atoms - 1)
        neighbors = build_neighbor_list(particle).get_neighbors(i)[0]
        if len(neighbors) == 0:
            continue
        j = random.choice(neighbors)
        if particle[i].symbol == particle[j].symbol:
            continue

        trial = particle.copy()
        trial[i].symbol, trial[j].symbol = trial[j].symbol, trial[i].symbol
        dE = calculate_energy(trial, coefficients, element_A) - energy

        accept = False
        if dE < 0:
            accept = True
        else:
            p_accept = math.exp(-dE / (boltzmann_k * temperature))
            if random.random() < p_accept:
                accept = True

        if accept:
            particle = trial
            energy += dE

        # Log and snapshot saving
        if step % save_interval == 0:
            total_A, surf, surf_A, ratio = count_surface(particle, element_A)
            log.append({
                'Step': step,
                'Energy (eV)': energy,
                f'Total {element_A}': total_A,
                'Surface Atoms': surf,
                f'{element_A} on Surface': surf_A,
                f'Surface {element_A} Ratio': ratio
            })
            status_text.text(
                f"Step {step} / {n_steps} | Energy: {energy:.4f} eV | Surface {element_A} Ratio: {ratio:.4f}"
            )

        if step % snapshot_interval == 0:
            snapshot_file = os.path.join(trajectory_dir, f"step_{step:05d}.xyz")
            write(snapshot_file, particle)

        progress_bar.progress(step / n_steps)

    elapsed = time.time() - start_time
    status_text.text(f"Simulation completed in {elapsed:.2f} seconds.")

    # Save final structure
    final_xyz_path = os.path.join(tempfile.gettempdir(), f"final_{element_A}{element_B}_{n_atoms}.xyz")
    write(final_xyz_path, particle)

    st.success(f"✅ Final structure saved: {final_xyz_path}")

    # Save log to Excel file
    if len(log) > 0:
        wb = Workbook()
        ws = wb.active
        ws.title = "Monte Carlo Log"

        # Parameters summary
        params = [
            ("Element A", element_A),
            ("Element B", element_B),
            ("Composition A", composition_A),
            ("Temperature (K)", temperature),
            ("MC Steps", n_steps),
            ("Save Interval", save_interval),
            ("Snapshot Interval", snapshot_interval),
            ("Lattice Type", lattice_type),
            ("Total Atoms", n_atoms)
        ]
        for i, (k, v) in enumerate(params, 1):
            ws.cell(row=i, column=1, value=k)
            ws.cell(row=i, column=2, value=v)

        # Header for log table
        header_row = len(params) + 2
        headers = list(log[0].keys())
        for col, header in enumerate(headers, start=1):
            ws.cell(row=header_row, column=col, value=header)

        # Log entries
        for row_idx, entry in enumerate(log, start=header_row + 1):
            for col_idx, key in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col_idx, value=entry[key])

        excel_path = os.path.join(tempfile.gettempdir(), f"MMC_{element_A}{element_B}_log.xlsx")
        wb.save(excel_path)
        st.success(f"📁 Excel log saved: {excel_path}")

        # Provide download button for Excel log
        with open(excel_path, "rb") as file:
            st.download_button(
                label="📥 Download Monte Carlo Log (Excel)",
                data=file,
                file_name=f"MMC_{element_A}{element_B}_log.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    else:
        st.warning("No log data was generated.")

    # Save final xyz file for download
    with open(final_xyz_path, "rb") as file:
        st.download_button(
            label="📥 Download Final Structure (XYZ file)",
            data=file,
            file_name=os.path.basename(final_xyz_path),
            mime="chemical/x-xyz"
        )

else:
    st.info("Run Monte Carlo simulation after generating nanoparticle in Phase 1.")
import matplotlib.pyplot as plt
from ase.io import read
from ase.visualize.plot import plot_atoms
from ase.data.colors import jmol_colors
import imageio.v2 as imageio
import openpyxl
import tempfile

st.markdown("---")
st.header("🎥 Phase 3: Visualization & Video Generation")

if 'particle' not in st.session_state or 'element_A' not in st.session_state:
    st.warning("Complete Phase 1 and Phase 2 before visualization.")
    st.stop()

trajectory_dir = os.path.join(tempfile.gettempdir(), "trajectory")
excel_path = os.path.join(tempfile.gettempdir(), f"MMC_{st.session_state['element_A']}{st.session_state['element_B']}_log.xlsx")

if not os.path.exists(trajectory_dir):
    st.warning("No trajectory folder found. Run Monte Carlo simulation in Phase 2 first.")
    st.stop()

if not os.path.exists(excel_path):
    st.warning("No Excel log file found. Run Monte Carlo simulation in Phase 2 first.")
    st.stop()

# Load Excel log data
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb.active

# Find header row after params (assumed 10th row by design)
header_row = 10
headers = [cell.value for cell in ws[header_row]]

step_col = headers.index('Step') + 1
energy_col = headers.index('Energy (eV)') + 1
surface_ratio_col = None
for key in headers:
    if key and 'Surface' in key and 'Ratio' in key:
        surface_ratio_col = headers.index(key) + 1
        break

steps = []
energies = []
surface_ratios = []

for row in ws.iter_rows(min_row=header_row+1, max_row=ws.max_row):
    steps.append(row[step_col - 1].value)
    energies.append(row[energy_col - 1].value)
    if surface_ratio_col:
        surface_ratios.append(row[surface_ratio_col - 1].value)
    else:
        surface_ratios.append(None)

# Find all snapshot xyz files
xyz_files = sorted([f for f in os.listdir(trajectory_dir) if f.endswith(".xyz")])
if len(xyz_files) == 0:
    st.warning("No snapshot files found in trajectory folder.")
    st.stop()

# Color map for elements - can extend if you add more
symbol_to_color = {
    st.session_state['element_A']: 'gold',
    st.session_state['element_B']: 'green',
    # Add more colors here if needed
}

def render_frame(xyz_path, step_label):
    atoms = read(xyz_path)

    colors = {}
    element_counts = {}
    for i, atom in enumerate(atoms):
        sym = atom.symbol
        colors[i] = symbol_to_color.get(sym, tuple(jmol_colors[atom.number]))
        element_counts[sym] = element_counts.get(sym, 0) + 1

    fig, ax = plt.subplots(figsize=(7,7), dpi=150)
    fig.patch.set_facecolor('#f0f0f0')
    ax.set_facecolor('#f0f0f0')

    plot_atoms(
        atoms, ax,
        radii=0.6,
        colors=colors,
        rotation=('30x,30y,0z'),
        show_unit_cell=False
    )

    # Frame label bottom-right
    fig.text(
        0.95, 0.05, step_label,
        ha='right', va='bottom',
        fontsize=10, color='black',
        bbox=dict(facecolor='white', edgecolor='gray', boxstyle='round,pad=0.3')
    )

    # Legend top-left with counts
    x0, y0 = 0.05, 0.95
    dy = 0.035
    for idx, sym in enumerate(sorted(element_counts.keys())):
        count = element_counts[sym]
        fig.text(
            x0, y0 - idx*dy,
            f"● {sym} ({count})",
            fontsize=9,
            ha='left', va='top',
            color=symbol_to_color.get(sym, 'gray'),
            fontweight='bold'
        )

    # Inset plot: energy & surface ratio up to current step
    inset_ax = fig.add_axes([0.6, 0.75, 0.35, 0.2])
    try:
        step_num = int(''.join(filter(str.isdigit, step_label)))
    except:
        step_num = 0
    idx_closest = min(range(len(steps)), key=lambda i: abs(steps[i] - step_num))

    inset_ax.plot(steps[:idx_closest+1], energies[:idx_closest+1], color='blue', label='Energy (eV)')
    if surface_ratio_col:
        inset_ax.plot(steps[:idx_closest+1], surface_ratios[:idx_closest+1], color='orange', label='Surface Ratio')

    inset_ax.set_xlabel('Step', fontsize=7)
    inset_ax.set_ylabel('Value', fontsize=7)
    inset_ax.tick_params(axis='both', labelsize=6)
    inset_ax.legend(fontsize=7)
    inset_ax.grid(True, linestyle='--', alpha=0.5)

    ax.axis('off')
    return fig

# Display frames one by one with option to select frame number
frame_idx = st.slider("Select snapshot frame", min_value=0, max_value=len(xyz_files)-1, value=0)
frame_file = xyz_files[frame_idx]
step_label = frame_file.replace(".xyz", "").replace("step_", "Step ")

st.pyplot(render_frame(os.path.join(trajectory_dir, frame_file), step_label))

# Option to generate video
if st.button("🎞️ Generate Evolution Video (MP4)"):

    fps = st.slider("Select video FPS (frames per second)", min_value=1, max_value=10, value=2)

    video_frames = []
    progress_bar = st.progress(0)
    status = st.empty()

    for i, xyz_file in enumerate(xyz_files):
        step_label = xyz_file.replace(".xyz", "").replace("step_", "Step ")
        fig = render_frame(os.path.join(trajectory_dir, xyz_file), step_label)

        # Save to PNG buffer
        import io
        buf = io.BytesIO()
        fig.savefig(buf, format='png')
        plt.close(fig)
        buf.seek(0)
        img = imageio.v2.imread(buf)
        video_frames.append(img)

        progress_bar.progress((i+1)/len(xyz_files))
        status.text(f"Rendering frame {i+1} of {len(xyz_files)}")

    # Save video to temp file
    video_file = os.path.join(tempfile.gettempdir(), "nanoparticle_evolution.mp4")
    imageio.mimsave(video_file, video_frames, fps=fps)

    st.success(f"✅ Video saved: {video_file}")
    with open(video_file, "rb") as f:
        st.download_button("📥 Download Video", f, file_name="nanoparticle_evolution.mp4", mime="video/mp4")

# Show plots of energy and surface ratio over all steps
st.subheader("📊 Energy and Surface Atom Ratio Evolution")

fig2, (ax1, ax2) = plt.subplots(1, 2, figsize=(12,4))
ax1.plot(steps, energies, color='blue', label='Energy (eV)')
ax1.set_xlabel('MC Step')
ax1.set_ylabel('Energy (eV)')
ax1.set_title('Energy vs. Step')
ax1.grid(True)

ax2.plot(steps, surface_ratios, color='orange', label='Surface Ratio')
ax2.set_xlabel('MC Step')
ax2.set_ylabel(f"{st.session_state['element_A']} Surface Atom Ratio")
ax2.set_title('Surface Atom Ratio vs. Step')
ax2.grid(True)

st.pyplot(fig2)
